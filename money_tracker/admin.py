from decimal import Decimal
from io import BytesIO

from django.contrib import admin
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone as django_timezone
from django.utils.html import format_html

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)

from .models import Account, Category, Transaction


admin.site.site_header = "Expense Admin"
admin.site.site_title = "Expense Admin"
admin.site.index_title = "Expense & Deposit Management"


@admin.register(Account)
class AccountAdmin(admin.ModelAdmin):
    list_display = [
        "name",
        "account_type",
        "opening_balance",
        "current_balance_display",
        "active",
    ]
    list_filter = ["account_type", "active"]
    search_fields = ["name"]

    def current_balance_display(self, obj):
        return obj.current_balance

    current_balance_display.short_description = "Current Balance"


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ["name", "transaction_type", "active"]
    list_filter = ["transaction_type", "active"]
    search_fields = ["name"]


@admin.register(Transaction)
class TransactionAdmin(admin.ModelAdmin):
    change_list_template = "admin/money_tracker/transaction/change_list.html"

    list_display = [
        "date",
        "transaction_type_badge",
        "account",
        "category",
        "amount",
        "created_by",
        "created_at",
    ]

    list_filter = [
        "transaction_type",
        "account",
        "category",
        "date",
    ]

    search_fields = [
        "account__name",
        "category__name",
        "note",
    ]

    date_hierarchy = "date"
    autocomplete_fields = ["account", "category"]
    readonly_fields = ["created_by", "created_at", "updated_at"]

    fieldsets = (
        ("Transaction Info", {
            "fields": (
                "transaction_type",
                "account",
                "category",
                "amount",
                "date",
            )
        }),
        ("Extra Info", {
            "fields": (
                "note",
                "attachment",
            )
        }),
        ("System Info", {
            "fields": (
                "created_by",
                "created_at",
                "updated_at",
            )
        }),
    )

    def transaction_type_badge(self, obj):
        if obj.transaction_type == "deposit":
            return format_html(
                '<span class="mt-badge mt-badge-deposit">{}</span>',
                "Deposit",
            )

        return format_html(
            '<span class="mt-badge mt-badge-expense">{}</span>',
            "Expense",
        )

    transaction_type_badge.short_description = "Type"

    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user

        obj.full_clean()
        super().save_model(request, obj, form, change)

    def get_urls(self):
        urls = super().get_urls()

        custom_urls = [
            path(
                "report/",
                self.admin_site.admin_view(self.report_view),
                name="money_tracker_transaction_report",
            ),
            path(
                "report/download/",
                self.admin_site.admin_view(self.report_download_view),
                name="money_tracker_transaction_report_download",
            ),
            path(
                "report/pdf/",
                self.admin_site.admin_view(self.report_pdf_view),
                name="money_tracker_transaction_report_pdf",
            ),
        ]

        return custom_urls + urls

    def get_report_queryset(self, request):
        transactions = Transaction.objects.select_related(
            "account",
            "category",
            "created_by",
        )

        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        account_id = request.GET.get("account")

        if start_date:
            transactions = transactions.filter(date__gte=start_date)

        if end_date:
            transactions = transactions.filter(date__lte=end_date)

        if account_id:
            transactions = transactions.filter(account_id=account_id)

        return transactions

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context=extra_context)

        try:
            qs = response.context_data["cl"].queryset

            totals = qs.aggregate(
                total_deposit=Sum(
                    "amount",
                    filter=Q(transaction_type="deposit"),
                ),
                total_expense=Sum(
                    "amount",
                    filter=Q(transaction_type="expense"),
                ),
            )

            total_deposit = totals["total_deposit"] or Decimal("0")
            total_expense = totals["total_expense"] or Decimal("0")

            response.context_data["total_deposit"] = total_deposit
            response.context_data["total_expense"] = total_expense
            response.context_data["net_balance"] = total_deposit - total_expense

        except Exception:
            pass

        return response

    def get_report_data(self, request):
        transactions = self.get_report_queryset(request)

        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        account_id = request.GET.get("account")

        totals = transactions.aggregate(
            total_deposit=Sum(
                "amount",
                filter=Q(transaction_type="deposit"),
            ),
            total_expense=Sum(
                "amount",
                filter=Q(transaction_type="expense"),
            ),
        )

        total_deposit = totals["total_deposit"] or Decimal("0")
        total_expense = totals["total_expense"] or Decimal("0")
        net_balance = total_deposit - total_expense

        daily_report_raw = transactions.values("date").annotate(
            deposit=Sum(
                "amount",
                filter=Q(transaction_type="deposit"),
            ),
            expense=Sum(
                "amount",
                filter=Q(transaction_type="expense"),
            ),
        ).order_by("-date")

        daily_report = []
        for row in daily_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            daily_report.append({
                "date": row["date"],
                "deposit": deposit,
                "expense": expense,
                "balance": deposit - expense,
            })

        account_report_raw = transactions.values("account__name").annotate(
            deposit=Sum(
                "amount",
                filter=Q(transaction_type="deposit"),
            ),
            expense=Sum(
                "amount",
                filter=Q(transaction_type="expense"),
            ),
        ).order_by("account__name")

        account_report = []
        for row in account_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            account_report.append({
                "account_name": row["account__name"],
                "deposit": deposit,
                "expense": expense,
                "balance": deposit - expense,
            })

        selected_account_name = "All Accounts"
        if account_id:
            account = Account.objects.filter(id=account_id).first()
            if account:
                selected_account_name = account.name

        return {
            "transactions": transactions,
            "transactions_ordered": transactions.order_by("-date", "-id"),
            "start_date": start_date,
            "end_date": end_date,
            "account_id": account_id,
            "selected_account_name": selected_account_name,
            "total_deposit": total_deposit,
            "total_expense": total_expense,
            "net_balance": net_balance,
            "daily_report": daily_report,
            "account_report": account_report,
            "daily_report_raw": daily_report_raw,
            "account_report_raw": account_report_raw,
        }

    def report_view(self, request):
        data = self.get_report_data(request)

        context = {
            **self.admin_site.each_context(request),
            "title": "Expense & Deposit Report",
            "transactions": data["transactions_ordered"],
            "accounts": Account.objects.filter(active=True),
            "start_date": data["start_date"],
            "end_date": data["end_date"],
            "selected_account": data["account_id"],
            "total_deposit": data["total_deposit"],
            "total_expense": data["total_expense"],
            "net_balance": data["net_balance"],
            "daily_report": data["daily_report"],
            "account_report": data["account_report"],
            "query_string": request.GET.urlencode(),
        }

        return TemplateResponse(
            request,
            "admin/money_tracker/transaction/report.html",
            context,
        )

    def report_download_view(self, request):
        data = self.get_report_data(request)

        transactions = data["transactions_ordered"]
        start_date = data["start_date"] or "All"
        end_date = data["end_date"] or "All"
        selected_account = data["selected_account_name"]
        total_deposit = data["total_deposit"]
        total_expense = data["total_expense"]
        net_balance = data["net_balance"]
        account_report_raw = data["account_report_raw"]
        daily_report_raw = data["daily_report_raw"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        primary_fill = PatternFill("solid", fgColor="0F766E")
        light_fill = PatternFill("solid", fgColor="F8FAFC")
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        deposit_fill = PatternFill("solid", fgColor="DCFCE7")
        expense_fill = PatternFill("solid", fgColor="FEE2E2")
        balance_fill = PatternFill("solid", fgColor="DBEAFE")
        white_fill = PatternFill("solid", fgColor="FFFFFF")

        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=18, bold=True, color="111827")
        subtitle_font = Font(size=10, color="6B7280")
        header_font = Font(bold=True, color="111827")
        bold_font = Font(bold=True, color="111827")

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        money_format = '#,##0.00'
        date_format = "mmm dd, yyyy"

        ws.merge_cells("A1:H1")
        ws["A1"] = "Expense & Deposit Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:H2")
        ws["A2"] = "Professional transaction summary generated from admin panel"
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A4"] = "Start Date"
        ws["B4"] = start_date
        ws["C4"] = "End Date"
        ws["D4"] = end_date
        ws["E4"] = "Account"
        ws["F4"] = selected_account

        for cell_ref in ["A4", "C4", "E4"]:
            ws[cell_ref].font = bold_font
            ws[cell_ref].fill = light_fill

        for cell_ref in ["A4", "B4", "C4", "D4", "E4", "F4"]:
            ws[cell_ref].border = thin_border
            ws[cell_ref].alignment = Alignment(vertical="center")

        ws.merge_cells("A6:B7")
        ws["A6"] = f"Total Deposit\n{total_deposit}"
        ws["A6"].fill = deposit_fill
        ws["A6"].font = Font(bold=True, size=12, color="166534")
        ws["A6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells("C6:D7")
        ws["C6"] = f"Total Expense\n{total_expense}"
        ws["C6"].fill = expense_fill
        ws["C6"].font = Font(bold=True, size=12, color="991B1B")
        ws["C6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells("E6:F7")
        ws["E6"] = f"Net Balance\n{net_balance}"
        ws["E6"].fill = balance_fill
        ws["E6"].font = Font(bold=True, size=12, color="1D4ED8")
        ws["E6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws["A6:F7"]:
            for cell in row:
                cell.border = thin_border

        current_row = 10

        def section_title(title, row):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1)
            cell.value = title
            cell.fill = primary_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            return row + 1

        def write_header(headers, row):
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            return row + 1

        current_row = section_title("Account Wise Report", current_row)
        current_row = write_header(["Account", "Deposit", "Expense", "Balance"], current_row)

        for row in account_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            balance = deposit - expense

            values = [
                row["account__name"],
                deposit,
                expense,
                balance,
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if col in [2, 3, 4]:
                    cell.number_format = money_format

                if col == 2:
                    cell.font = Font(bold=True, color="166534")
                elif col == 3:
                    cell.font = Font(bold=True, color="991B1B")
                elif col == 4:
                    cell.font = Font(bold=True, color="1D4ED8")

            current_row += 1

        current_row += 2

        current_row = section_title("Daily Report", current_row)
        current_row = write_header(["Date", "Deposit", "Expense", "Balance"], current_row)

        for row in daily_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            balance = deposit - expense

            values = [
                row["date"],
                deposit,
                expense,
                balance,
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if col == 1:
                    cell.number_format = date_format
                elif col in [2, 3, 4]:
                    cell.number_format = money_format

                if col == 2:
                    cell.font = Font(bold=True, color="166534")
                elif col == 3:
                    cell.font = Font(bold=True, color="991B1B")
                elif col == 4:
                    cell.font = Font(bold=True, color="1D4ED8")

            current_row += 1

        current_row += 2

        current_row = section_title("Transaction Details", current_row)
        transaction_header_row = current_row

        current_row = write_header(
            [
                "Date",
                "Type",
                "Account",
                "Category",
                "Amount",
                "Note",
                "Created By",
                "Created At",
            ],
            current_row,
        )

        for transaction in transactions:
            created_at = ""
            if transaction.created_at:
                created_at = django_timezone.localtime(transaction.created_at).replace(tzinfo=None)

            values = [
                transaction.date,
                transaction.get_transaction_type_display(),
                transaction.account.name,
                str(transaction.category),
                transaction.amount,
                transaction.note or "",
                transaction.created_by.username if transaction.created_by else "",
                created_at,
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if col == 1:
                    cell.number_format = date_format

                elif col == 2:
                    if transaction.transaction_type == "deposit":
                        cell.fill = deposit_fill
                        cell.font = Font(bold=True, color="166534")
                    else:
                        cell.fill = expense_fill
                        cell.font = Font(bold=True, color="991B1B")

                elif col == 5:
                    cell.number_format = money_format

                    if transaction.transaction_type == "deposit":
                        cell.font = Font(bold=True, color="166534")
                    else:
                        cell.font = Font(bold=True, color="991B1B")

                elif col == 8 and created_at:
                    cell.number_format = "mmm dd, yyyy h:mm AM/PM"

            current_row += 1

        if not transactions.exists():
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = ws.cell(row=current_row, column=1)
            cell.value = "No transaction found for selected filter."
            cell.fill = white_fill
            cell.font = Font(color="6B7280", italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

        ws.freeze_panes = f"A{transaction_header_row + 1}"

        if current_row > transaction_header_row:
            ws.auto_filter.ref = f"A{transaction_header_row}:H{current_row - 1}"

        widths = {
            "A": 16,
            "B": 14,
            "C": 20,
            "D": 28,
            "E": 16,
            "F": 40,
            "G": 18,
            "H": 24,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 30

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="expense_deposit_report.xlsx"'

        wb.save(response)
        return response

    def report_pdf_view(self, request):
        data = self.get_report_data(request)

        transactions = data["transactions_ordered"]
        start_date = data["start_date"] or "All"
        end_date = data["end_date"] or "All"
        selected_account = data["selected_account_name"]
        total_deposit = data["total_deposit"]
        total_expense = data["total_expense"]
        net_balance = data["net_balance"]
        account_report_raw = data["account_report_raw"]
        daily_report_raw = data["daily_report_raw"]

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=28,
            leftMargin=28,
            topMargin=34,
            bottomMargin=28,
            title="Expense & Deposit Report",
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "SmartTitle",
            parent=styles["Title"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#111827"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )

        subtitle_style = ParagraphStyle(
            "SmartSubtitle",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#6B7280"),
            alignment=TA_CENTER,
            spaceAfter=10,
        )

        section_style = ParagraphStyle(
            "SectionTitle",
            parent=styles["Heading2"],
            fontSize=11,
            leading=14,
            textColor=colors.white,
            spaceBefore=0,
            spaceAfter=0,
        )

        normal_style = ParagraphStyle(
            "NormalSmall",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#111827"),
        )

        def money(value):
            value = value or Decimal("0")
            return f"{value:,.2f}"

        def safe_text(value):
            if value is None:
                return ""
            return str(value)

        def footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#6B7280"))
            canvas.drawString(28, 18, "Expense Admin - Expense & Deposit Report")
            canvas.drawRightString(
                landscape(A4)[0] - 28,
                18,
                f"Page {doc_obj.page}",
            )
            canvas.restoreState()

        def section_header(title):
            table = Table(
                [[Paragraph(title, section_style)]],
                colWidths=[doc.width],
            )
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F766E")),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.HexColor("#0F766E")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            return table

        def smart_table(data_rows, widths=None, header=True):
            table = Table(
                data_rows,
                colWidths=widths,
                repeatRows=1 if header else 0,
            )

            style = [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]

            if header:
                style += [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ]

            table.setStyle(TableStyle(style))
            return table

        elements = []

        elements.append(Paragraph("Expense & Deposit Report", title_style))
        elements.append(Paragraph("Smart PDF report generated from admin panel", subtitle_style))

        filter_data = [
            [
                Paragraph("<b>Start Date</b>", normal_style),
                Paragraph(safe_text(start_date), normal_style),
                Paragraph("<b>End Date</b>", normal_style),
                Paragraph(safe_text(end_date), normal_style),
                Paragraph("<b>Account</b>", normal_style),
                Paragraph(safe_text(selected_account), normal_style),
            ]
        ]

        filter_table = smart_table(
            filter_data,
            widths=[80, 120, 80, 120, 70, 180],
            header=False,
        )
        filter_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ]))

        elements.append(filter_table)
        elements.append(Spacer(1, 8))

        summary_data = [
            [
                Paragraph(f"<b>Total Deposit</b><br/>{money(total_deposit)}", normal_style),
                Paragraph(f"<b>Total Expense</b><br/>{money(total_expense)}", normal_style),
                Paragraph(f"<b>Net Balance</b><br/>{money(net_balance)}", normal_style),
            ]
        ]

        summary_table = Table(
            summary_data,
            colWidths=[doc.width / 3 - 5, doc.width / 3 - 5, doc.width / 3 - 5],
        )

        summary_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#DCFCE7")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#FEE2E2")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#DBEAFE")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))

        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        elements.append(section_header("Account Wise Report"))

        account_data = [
            ["Account", "Deposit", "Expense", "Balance"]
        ]

        for row in account_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            balance = deposit - expense

            account_data.append([
                safe_text(row["account__name"]),
                money(deposit),
                money(expense),
                money(balance),
            ])

        if len(account_data) == 1:
            account_data.append(["No data found", "", "", ""])

        account_table = smart_table(
            account_data,
            widths=[260, 130, 130, 130],
        )

        account_table.setStyle(TableStyle([
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("TEXTCOLOR", (1, 1), (1, -1), colors.HexColor("#166534")),
            ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#991B1B")),
            ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#1D4ED8")),
        ]))

        elements.append(account_table)
        elements.append(Spacer(1, 10))

        elements.append(section_header("Daily Report"))

        daily_data = [
            ["Date", "Deposit", "Expense", "Balance"]
        ]

        for row in daily_report_raw:
            deposit = row["deposit"] or Decimal("0")
            expense = row["expense"] or Decimal("0")
            balance = deposit - expense

            daily_data.append([
                row["date"].strftime("%d %b %Y") if row["date"] else "",
                money(deposit),
                money(expense),
                money(balance),
            ])

        if len(daily_data) == 1:
            daily_data.append(["No data found", "", "", ""])

        daily_table = smart_table(
            daily_data,
            widths=[160, 150, 150, 150],
        )

        daily_table.setStyle(TableStyle([
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("TEXTCOLOR", (1, 1), (1, -1), colors.HexColor("#166534")),
            ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#991B1B")),
            ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#1D4ED8")),
        ]))

        elements.append(daily_table)
        elements.append(PageBreak())

        elements.append(Paragraph("Transaction Details", title_style))
        elements.append(Paragraph("All matching transactions from selected filter", subtitle_style))

        transaction_data = [
            [
                "Date",
                "Type",
                "Account",
                "Category",
                "Amount",
                "Note",
                "Created By",
            ]
        ]

        for transaction in transactions:
            transaction_data.append([
                transaction.date.strftime("%d %b %Y") if transaction.date else "",
                transaction.get_transaction_type_display(),
                safe_text(transaction.account.name),
                safe_text(transaction.category.name),
                money(transaction.amount),
                Paragraph(safe_text(transaction.note or "-"), normal_style),
                transaction.created_by.username if transaction.created_by else "",
            ])

        if len(transaction_data) == 1:
            transaction_data.append(["No transaction found", "", "", "", "", "", ""])

        transaction_table = smart_table(
            transaction_data,
            widths=[75, 65, 110, 130, 85, 250, 90],
        )

        transaction_table.setStyle(TableStyle([
            ("ALIGN", (4, 1), (4, -1), "RIGHT"),
            ("TEXTCOLOR", (4, 1), (4, -1), colors.HexColor("#111827")),
        ]))

        elements.append(transaction_table)

        doc.build(
            elements,
            onFirstPage=footer,
            onLaterPages=footer,
        )

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="expense_deposit_report.pdf"'
        response.write(pdf)

        return response